import streamlit as st
import pandas as pd
from io import BytesIO
import re
import openpyxl
import tempfile
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import pymongo
import seaborn as sns
import matplotlib.pyplot as plt
import streamlit_dynamic_filters as sdf
import hmac

st.set_page_config(
    page_title="Staff Production Management",
    page_icon="hammer_and_wrench",
    layout="wide"
)

st.title('Staff Production Management')
sns.set_theme(style="darkgrid")

def check_password():
    """Returns `True` if the user had the correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if hmac.compare_digest(st.session_state["password"], st.secrets["app_pwd"]):
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # Don't store the password.
        else:
            st.session_state["password_correct"] = False

    # Return True if the password is validated.
    if st.session_state.get("password_correct", False):
        return True

    # Show input for password.
    st.text_input(
        "Password", type="password", on_change=password_entered, key="password"
    )
    if "password_correct" in st.session_state:
        st.error("😕 Password incorrect")
    return False


if not check_password():
    st.stop()  # Do not continue if check_password is not True.

@st.cache_resource
def init_connection():
    return pymongo.MongoClient(st.secrets["db_cs"])

client = init_connection()

@st.cache_data(ttl=600)
def get_procedures():
    db = client["staff_db"]
    coll = db["procedures"]
    return list(coll.find())

@st.cache_data(ttl=600)
def get_main_performances():
    db = client["staff_db"]
    coll = db["main_performances"]
    return list(coll.find())

@st.cache_data(ttl=600)
def get_available_productions_for_ivs():
    db = client["staff_db"]
    coll = db["iv"]
    return coll.distinct("production_ref")

@st.cache_data(ttl=600)
def get_ivs_for_production_ref(production_ref):
    db = client["staff_db"]
    coll = db["iv"]
    return list(coll.find({"production_ref": production_ref}))

tab1,tab2,tab3 = st.tabs(["Load and View Procedure Catalog","Load Init File and Create Logbook","Dashboards"])

with tab1:
    if st.button("Load procedures catalog from db"):
        df = pd.DataFrame(columns=["procedure_name","procedure_version","linked_block","data_name","data_description","recipe_value","data_type","data_unit","data_min_value","data_max_value","data_origin","data_perimeter"])
        for doc in get_procedures():
            new_df = pd.DataFrame(columns=["procedure_name","procedure_version","linked_block","data_name","data_description","recipe_value","data_type","data_unit","data_min_value","data_max_value","data_origin","data_perimeter"])
            for data in doc["procedure_data"]:
                new_row = {"procedure_name":doc["procedure_name"],"procedure_version":doc["procedure_version"],"linked_block":doc["linked_block"],"data_name":data["data_name"],"data_description":data["data_description"],"recipe_value":data["recipe_value"],"data_type":data["data_type"],"data_unit":data["data_unit"],"data_min_value":data["data_min_value"],"data_max_value":data["data_max_value"],"data_origin":data["data_origin"],"data_perimeter":data["data_perimeter"]}
                new_df = pd.concat([new_df,pd.DataFrame([new_row])],axis=0)
            df = pd.concat([df,new_df],axis=0)
        st.session_state["df"] = df
    if "df" in st.session_state:
        st.success("Procedures catalog loaded.")
        dynamic_filters = sdf.DynamicFilters(df=st.session_state["df"], filters=['procedure_name', 'procedure_version', 'linked_block'])
        dynamic_filters.display_filters(location="columns",num_columns=2,gap="small")
        dynamic_filters.display_df()
with tab2:
    if "df" in st.session_state:
        st.success("Procedures catalog loaded.")
        uploaded_init = st.file_uploader("Select the production initialization file to create the logbook from.", type="xlsx")
        if uploaded_init:
            init_df = pd.read_excel(uploaded_init)
            production_ref = re.findall(r'(?i)ST\d{2}',uploaded_init.name)[0].upper()
            st.write(f"Production reference : {production_ref}")
            st.write(f"{init_df["stack_ref"].unique().size} stacks in the initialization file : {init_df["stack_ref"].unique()}")
            valid_init = True
            df = st.session_state["df"]
            
            for procedure_name,procedure_version in zip(init_df["procedure_name"],init_df["procedure_version"]):      
                if not ((df["procedure_name"]==procedure_name) & (df["procedure_version"]==procedure_version)).any():
                    st.write(f"Procedure {procedure_name} v{procedure_version} not found in the catalog. Please add it in the catalog and save cache changes or remove it from the initialization file before proceeding.")
                    valid_init = False
            if valid_init:
                st.success("Initialization file is valid.")
                bg_color = st.color_picker("Pick a color for the locked cells in the logbook (default : black)")
                bg_color = bg_color[1:]
                if st.button("Create logbook"):

                    st.write("Creating logbook ...")
                    MAX_COLS = 20
                    full_logbook = {}
                    wb = openpyxl.Workbook()
                    production_data = df.loc[df["data_type"]=="production"]
                    for procedure_name,procedure_version,stack_ref in zip(init_df["procedure_name"],init_df["procedure_version"],init_df["stack_ref"]):
                        target_production_data = production_data.loc[(production_data["procedure_name"]==procedure_name) & (production_data["procedure_version"]==procedure_version)]
                        if target_production_data.empty:
                            continue
                        new_logbook_chunk = pd.DataFrame(columns=["stack_ref","procedure_name","procedure_version","data_name","data_description","data_unit","batch_data"]+[f"run_{i}" for i in range(1,MAX_COLS+1)])
                        for data in target_production_data.iterrows():
                            # data[0] is the df index, data[1] is the row
                            # for front-end purposes, we'll mark the soon-to-be greyed out cells with an "x" and format later
                            # WARNING : if columns are added here (before data_unit), the lock/unlock will need to be updated (data_unit won't be E column anymore)
                            new_row = {"stack_ref":stack_ref,"procedure_name":procedure_name,"procedure_version":procedure_version,"data_name":data[1]["data_name"],"data_description":data[1]["data_description"],"data_unit":data[1]["data_unit"],"batch_data": "x" if data[1]["data_perimeter"] == "run" else ""}
                            if data[1]["data_perimeter"] == "run":
                                for i in range(1,MAX_COLS+1):
                                    new_row.update({f"run_{i}": ""})
                            else:
                                for i in range(1,MAX_COLS+1):
                                    new_row.update({f"run_{i}": "x"})
                            new_logbook_chunk = pd.concat([new_logbook_chunk,pd.DataFrame([new_row])],axis=0)
                        key = target_production_data["linked_block"].iloc[0] # there should only be one ...
                        if key in full_logbook:
                            full_logbook[key] = pd.concat([full_logbook[key],new_logbook_chunk],axis=0)
                        else:
                            full_logbook[key] = new_logbook_chunk

                    for k,v in full_logbook.items():
                        # group same version of the same procedure together in the chunks, aggregate stacks in stack_ref
                        v = v.groupby([col for col in v.columns if col != "stack_ref"], dropna=False,sort=False)["stack_ref"].apply(', '.join).reset_index()
                        # reorder columns
                        cols = v.columns.tolist()
                        cols = cols[-1:] + cols[:-1]
                        v = v[cols]

                        wb.create_sheet(title=k)
                        ws = wb[k]
                        for r in dataframe_to_rows(v, index=False, header=True):
                            ws.append(r)
                    wb.remove(wb["Sheet"])

                    font = Font(bold=True)

                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        ws.protection.sheet = True
                        for cell in ws["1:1"]:
                            cell.font = font

                        # Iterate over all columns and set the width to the maximum length of the cell content in each column
                        # Greys out cells marked with "x" and empty them
                        for col in ws.columns:
                            max_length = 0
                            column = col[0].column_letter  # Get the column name
                            for cell in col:
                                if not cell.value and column != "F":
                                    # F column is data_unit, special case
                                    # if the cell was empty and was not in the unit column, we make it editable
                                    cell.protection = openpyxl.styles.Protection(locked=False)            
                                elif cell.value == "x":
                                    cell.value = ""
                                    cell.fill = openpyxl.styles.PatternFill(fgColor=bg_color, fill_type = "solid")
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            ws.column_dimensions[column].width = adjusted_width

                    data = BytesIO()
                    wb.save(data)
                    
                    st.download_button("Download logbook",data=data,mime='xlsx',file_name=f"{production_ref}_logbook.xlsx")


                    
    else:
        st.warning("No procedures catalog available, fetch them from db through previous tab.")
with tab3:

    st.header("Main Performances")

    st.write("Select a main performance to display")
    selected_main_performance = st.selectbox("Available main performances : ", ["voc_be","jsc_be","ff_be","pce_be","voc_ae","jsc_ae","ff_ae","pce_ae"])

    def kde_rug_plot_main_performance(selected_main_performance,selected_bw_adjust):
        df = pd.DataFrame(get_main_performances())
        df.drop(columns=["_id"], inplace=True)

        fig = plt.figure(figsize=(10, 4))
    
        sns.kdeplot(data=df, x=selected_main_performance, hue="production_ref",bw_adjust=selected_bw_adjust).set_title(f"KDE plot of {selected_main_performance} across productions")
        sns.rugplot(data=df, x=selected_main_performance, hue="production_ref", height=-0.03, clip_on=False)
        st.pyplot(fig,use_container_width=False)

    if selected_main_performance:
        selected_bw_adjust = st.slider("Bandwidth adjustment", min_value=0.1, max_value=5.0, value=1.0, step=0.1)
        if st.button("Query and plot selected main performance"):
            kde_rug_plot_main_performance(selected_main_performance,selected_bw_adjust)

    st.divider()

    st.header("IVs")

    st.write("Select a production reference to display its IVs")
    selected_production_for_ivs = st.selectbox("Available productions : ", get_available_productions_for_ivs())

    if selected_production_for_ivs:
        if st.button("Query and plot IVs"):

            data = get_ivs_for_production_ref(selected_production_for_ivs)
            df = pd.DataFrame(columns=["device_ref","voltage_fw","current_density_fw","power_fw","voltage_rv","current_density_rv","power_rv"])
            for d in data:
                del d["_id"]
                device_ref = d["device_ref"]
                for a,b,c,d,e,f in zip(d["iv_data"]["voltage_fw"],d["iv_data"]["current_density_fw"],d["iv_data"]["power_fw"],d["iv_data"]["voltage_rv"],d["iv_data"]["current_density_rv"],d["iv_data"]["power_rv"]):
                    new_row = {"device_ref":device_ref,"voltage_fw":a,"current_density_fw":b,"power_fw":c,"voltage_rv":d,"current_density_rv":e,"power_rv":f}
                    df = pd.concat([df,pd.DataFrame([new_row])],axis=0,ignore_index=True)
            
            fig = plt.figure(figsize=(12, 6))
            plt.subplot(2,1,1)
            sns.lineplot(data=df, x="voltage_fw", y="current_density_fw", hue="device_ref").set_title(f"{selected_production_for_ivs} : IV Forward")
            plt.subplot(2,1,2)
            sns.lineplot(data=df, x="voltage_rv", y="current_density_rv", hue="device_ref").set_title(f"{selected_production_for_ivs} : IV Reverse")
            fig.tight_layout()
            st.pyplot(fig,use_container_width=False)







            